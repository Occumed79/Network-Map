#!/usr/bin/env python3

from __future__ import annotations

import json
import math
import os
import re
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import psycopg
import requests
from openpyxl import load_workbook

OUTPUT_DIR = Path("data/generated/who-hidr-healthcare")
STATUS_DIR = Path("data/status")
SOURCE_NAME = "WHO Health Inequality Data Repository"
SOURCE_PAGE = "https://www.who.int/data/inequality-monitor/data"

DATASETS = {
    "rep_gho_hc": "https://datasafe-h5afbhf4gwctabaa.z01.azurefd.net/api/Download/TOP/rep_gho_hc/data",
    "rep_dhs_hca": "https://datasafe-h5afbhf4gwctabaa.z01.azurefd.net/api/Download/TOP/rep_dhs_hca/data",
    "rep_oecd_hc": "https://datasafe-h5afbhf4gwctabaa.z01.azurefd.net/api/Download/TOP/rep_oecd_hc/data",
    "rep_eurostat_hc": "https://datasafe-h5afbhf4gwctabaa.z01.azurefd.net/api/Download/TOP/rep_eurostat_hc/data",
}

DB_ENV_NAMES = (
    "DATABASE_URL",
    "NEON_DATABASE_URL",
    "POSTGRES_URL",
    "POSTGRES_DATABASE_URL",
)


def clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def numeric(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    return parsed if math.isfinite(parsed) else None


def integer_year(value: Any) -> int | None:
    number = numeric(value)
    if number is None:
        return None
    year = int(number)
    return year if 1900 <= year <= 2200 else None


def parse_updated_at(value: Any) -> datetime | None:
    text = clean(value)
    if not text:
        return None
    for fmt in ("%d %B %Y", "%d %b %Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def infer_unit(name: str, scale: Any) -> str:
    lowered = name.lower()
    if "(%)" in name or "percent" in lowered or "percentage" in lowered:
        return "%"
    match = re.search(r"per\s+([\d\s,]+(?:patients|population|people|births)?)", lowered)
    if match:
        return f"per {clean(match.group(1))}"
    scale_value = numeric(scale)
    if scale_value == 100:
        return "0-100 scale"
    return "value"


def geography(dataset_id: str, dimension: str, subgroup: str, reg_id: str) -> tuple[str, str | None, str | None]:
    dimension_lower = dimension.lower()
    if dataset_id == "rep_dhs_hca" and reg_id and ("subnational" in dimension_lower or "region" in dimension_lower):
        return "admin1", reg_id, subgroup or None
    return "country", None, None


def get_database_url() -> tuple[str, str]:
    for name in DB_ENV_NAMES:
        value = os.getenv(name)
        if value:
            return value, name
    raise RuntimeError(f"No database URL found in any of: {', '.join(DB_ENV_NAMES)}")


def download_workbooks() -> dict[str, Path]:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    session = requests.Session()
    session.headers["User-Agent"] = "Occu-Med-Network-Map/1.0 WHO-HIDR-healthcare-import"
    paths: dict[str, Path] = {}
    for dataset_id, url in DATASETS.items():
        response = session.get(url, timeout=240)
        response.raise_for_status()
        path = OUTPUT_DIR / f"{dataset_id}.xlsx"
        path.write_bytes(response.content)
        paths[dataset_id] = path
    return paths


def rows_from_workbook(dataset_id: str, path: Path, diagnostics: Counter[str]) -> Iterable[tuple[Any, ...]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Data" not in workbook.sheetnames:
        raise RuntimeError(f"{dataset_id}: workbook has no Data sheet")
    sheet = workbook["Data"]
    iterator = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(iterator)]
    required = {"setting", "date", "source", "indicator_abbr", "indicator_name", "dimension", "subgroup", "estimate", "iso3", "dataset_id"}
    missing = sorted(required.difference(headers))
    if missing:
        raise RuntimeError(f"{dataset_id}: missing required columns {missing}")

    for row_number, values in enumerate(iterator, start=2):
        row = dict(zip(headers, values))
        estimate = numeric(row.get("estimate"))
        if estimate is None:
            diagnostics["missing_or_invalid_estimate"] += 1
            continue

        year = integer_year(row.get("date"))
        if year is None:
            diagnostics["invalid_year"] += 1
            continue

        country_code = clean(row.get("iso3")).upper()
        if not re.fullmatch(r"[A-Z]{3}", country_code):
            diagnostics["invalid_iso3"] += 1
            continue

        indicator_code = clean(row.get("indicator_abbr"))
        if not indicator_code:
            diagnostics["missing_indicator_code"] += 1
            continue

        dimension = clean(row.get("dimension")) or "All"
        subgroup = clean(row.get("subgroup")) or "All"
        reg_id = clean(row.get("reg_id"))
        geography_level, admin1_code, admin1_name = geography(dataset_id, dimension, subgroup, reg_id)

        lower = numeric(row.get("ci_lb"))
        upper = numeric(row.get("ci_ub"))
        if lower is not None and upper is not None and lower > upper:
            diagnostics["invalid_confidence_interval"] += 1
            lower = None
            upper = None

        dimension_payload = {
            "dimension": dimension,
            "subgroup": subgroup,
        }
        if reg_id:
            dimension_payload["reg_id"] = reg_id
        dimension_key = json.dumps(dimension_payload, sort_keys=True, ensure_ascii=False, separators=(",", ":"))

        metadata = {
            "upstream_source": clean(row.get("source")),
            "standard_error": numeric(row.get("se")),
            "population": numeric(row.get("population")),
            "flag": clean(row.get("flag")) or None,
            "setting_average": numeric(row.get("setting_average")),
            "favourable_indicator": row.get("favourable_indicator"),
            "indicator_scale": numeric(row.get("indicator_scale")),
            "ordered_dimension": row.get("ordered_dimension"),
            "subgroup_order": row.get("subgroup_order"),
            "reference_subgroup": row.get("reference_subgroup"),
            "who_region": clean(row.get("whoreg6")) or None,
            "world_bank_income_group": clean(row.get("wbincome2025")) or None,
            "source_update_text": clean(row.get("update")) or None,
            "source_row_number": row_number,
            "missing_values_omitted": True,
            "frontend_storage": False,
        }

        indicator_name = clean(row.get("indicator_name"))
        source_url = DATASETS[dataset_id]

        yield (
            dataset_id,
            indicator_code,
            indicator_name or None,
            country_code,
            clean(row.get("setting")) or None,
            admin1_code,
            admin1_name,
            geography_level,
            year,
            dimension_key,
            json.dumps(dimension_payload, ensure_ascii=False, separators=(",", ":")),
            subgroup,
            estimate,
            infer_unit(indicator_name, row.get("indicator_scale")),
            "subgroup_estimate",
            lower,
            upper,
            SOURCE_NAME,
            source_url,
            parse_updated_at(row.get("update")),
            json.dumps(metadata, ensure_ascii=False, separators=(",", ":")),
        )


STAGE_COLUMNS = (
    "dataset_id",
    "indicator_code",
    "indicator_name",
    "country_code",
    "country_name",
    "admin1_code",
    "admin1_name",
    "geography_level",
    "year",
    "dimension_key",
    "dimensions",
    "subgroup_label",
    "value",
    "unit",
    "measure_type",
    "lower_bound",
    "upper_bound",
    "source_name",
    "source_url",
    "source_updated_at",
    "metadata",
)


def create_stage(cursor: psycopg.Cursor[Any]) -> None:
    cursor.execute(
        """
        CREATE TEMP TABLE IF NOT EXISTS stage_who_hidr_healthcare (
          dataset_id text NOT NULL,
          indicator_code text NOT NULL,
          indicator_name text,
          country_code text NOT NULL,
          country_name text,
          admin1_code text,
          admin1_name text,
          geography_level text NOT NULL,
          year integer NOT NULL,
          dimension_key text NOT NULL,
          dimensions jsonb NOT NULL,
          subgroup_label text,
          value numeric NOT NULL,
          unit text,
          measure_type text,
          lower_bound numeric,
          upper_bound numeric,
          source_name text NOT NULL,
          source_url text,
          source_updated_at timestamptz,
          metadata jsonb NOT NULL
        ) ON COMMIT PRESERVE ROWS
        """
    )


def load_dataset(connection: psycopg.Connection[Any], dataset_id: str, path: Path) -> dict[str, Any]:
    diagnostics: Counter[str] = Counter()
    copied = 0
    with connection.cursor() as cursor:
        create_stage(cursor)
        cursor.execute("TRUNCATE stage_who_hidr_healthcare")
        columns = ", ".join(STAGE_COLUMNS)
        with cursor.copy(f"COPY stage_who_hidr_healthcare ({columns}) FROM STDIN") as copy:
            for record in rows_from_workbook(dataset_id, path, diagnostics):
                copy.write_row(record)
                copied += 1

        cursor.execute(
            """
            INSERT INTO public.international_health_inequality_observations (
              dataset_id, indicator_code, indicator_name, country_code, country_name,
              admin1_code, admin1_name, geography_level, year, dimension_key,
              dimensions, subgroup_label, value, unit, measure_type, lower_bound,
              upper_bound, source_name, source_url, source_updated_at, metadata
            )
            SELECT
              dataset_id, indicator_code, indicator_name, country_code, country_name,
              admin1_code, admin1_name, geography_level, year, dimension_key,
              dimensions, subgroup_label, value, unit, measure_type, lower_bound,
              upper_bound, source_name, source_url, source_updated_at, metadata
            FROM stage_who_hidr_healthcare
            ON CONFLICT (
              dataset_id,
              indicator_code,
              country_code,
              (COALESCE(admin1_code, ''::text)),
              geography_level,
              year,
              dimension_key,
              source_name
            ) DO UPDATE SET
              indicator_name = EXCLUDED.indicator_name,
              country_name = EXCLUDED.country_name,
              admin1_name = EXCLUDED.admin1_name,
              dimensions = EXCLUDED.dimensions,
              subgroup_label = EXCLUDED.subgroup_label,
              value = EXCLUDED.value,
              unit = EXCLUDED.unit,
              measure_type = EXCLUDED.measure_type,
              lower_bound = EXCLUDED.lower_bound,
              upper_bound = EXCLUDED.upper_bound,
              source_url = EXCLUDED.source_url,
              source_updated_at = EXCLUDED.source_updated_at,
              metadata = EXCLUDED.metadata,
              updated_at = now()
            """
        )
        cursor.execute(
            """
            SELECT
              COUNT(*)::int AS rows,
              COUNT(DISTINCT country_code)::int AS countries,
              COUNT(DISTINCT indicator_code)::int AS indicators,
              MIN(year)::int AS oldest_year,
              MAX(year)::int AS newest_year,
              COUNT(*) FILTER (WHERE geography_level = 'admin1')::int AS admin1_rows
            FROM public.international_health_inequality_observations
            WHERE dataset_id = %s
            """,
            (dataset_id,),
        )
        counts = cursor.fetchone()
    connection.commit()
    return {
        "dataset_id": dataset_id,
        "copied_rows": copied,
        "stored_rows": counts[0],
        "countries": counts[1],
        "indicators": counts[2],
        "oldest_year": counts[3],
        "newest_year": counts[4],
        "admin1_rows": counts[5],
        "skipped": dict(diagnostics),
        "source_url": DATASETS[dataset_id],
    }


def verify_database(connection: psycopg.Connection[Any]) -> None:
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT
              current_database(),
              to_regclass('public.medical_providers')::text,
              to_regclass('public.international_health_indicators')::text,
              to_regclass('public.international_health_inequality_observations')::text
            """
        )
        database_name, medical_table, indicator_table, inequality_table = cursor.fetchone()
    expected = ("neondb", "medical_providers", "international_health_indicators", "international_health_inequality_observations")
    actual = (database_name, medical_table, indicator_table, inequality_table)
    if actual != expected:
        raise RuntimeError(f"Database identity rejected: expected {expected}, received {actual}")


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    STATUS_DIR.mkdir(parents=True, exist_ok=True)
    database_url, secret_name = get_database_url()
    workbook_paths = download_workbooks()
    results = []

    with psycopg.connect(database_url, connect_timeout=30) as connection:
        verify_database(connection)
        for dataset_id, path in workbook_paths.items():
            results.append(load_dataset(connection, dataset_id, path))

        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT
                  COUNT(*)::int,
                  COUNT(DISTINCT dataset_id)::int,
                  COUNT(DISTINCT country_code)::int,
                  COUNT(DISTINCT indicator_code)::int,
                  COUNT(*) FILTER (WHERE value IS NULL)::int,
                  COUNT(*) FILTER (WHERE geography_level = 'admin1')::int
                FROM public.international_health_inequality_observations
                WHERE dataset_id = ANY(%s)
                """,
                (list(DATASETS),),
            )
            totals = cursor.fetchone()

    manifest = {
        "status": "IMPORTED",
        "source": SOURCE_NAME,
        "source_page": SOURCE_PAGE,
        "database": "Map / neondb / production",
        "database_secret_name": secret_name,
        "frontend_storage": False,
        "datasets": results,
        "totals": {
            "rows": totals[0],
            "datasets": totals[1],
            "countries": totals[2],
            "indicators": totals[3],
            "null_values": totals[4],
            "admin1_rows": totals[5],
        },
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "safeguards": [
            "Official WHO HIDR structured downloads only",
            "Missing or invalid estimates omitted",
            "No frontend data bundle",
            "Database identity verified before writes",
            "Idempotent production upsert",
        ],
    }
    (OUTPUT_DIR / "import-summary.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    (STATUS_DIR / "who-hidr-healthcare-import-summary.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    (STATUS_DIR / "who-hidr-healthcare-import-status.txt").write_text("IMPORTED\n", encoding="utf-8")
    print(json.dumps(manifest, indent=2))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        STATUS_DIR.mkdir(parents=True, exist_ok=True)
        error = {"status": "FAILED", "error": str(exc), "generated_at": datetime.now(timezone.utc).isoformat()}
        (OUTPUT_DIR / "import-summary.json").write_text(json.dumps(error, indent=2), encoding="utf-8")
        (STATUS_DIR / "who-hidr-healthcare-import-summary.json").write_text(json.dumps(error, indent=2), encoding="utf-8")
        (STATUS_DIR / "who-hidr-healthcare-import-status.txt").write_text("FAILED\n", encoding="utf-8")
        print(json.dumps(error, indent=2), file=sys.stderr)
        raise
