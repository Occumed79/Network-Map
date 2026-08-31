#!/usr/bin/env python3
import argparse, csv, hashlib, json, re, unicodedata
from pathlib import Path

COLUMNS = ["source_record_id","source_url","name","normalized_name","address_line1","formatted_address","city","state_region","postal_code","country_code","lat","lng","phone","website","email","primary_provider_type","capability_tags","quality_score","master_key"]
DATASET_URL = "https://datos.salud.gob.ar/dataset/listado-establecimientos-de-salud-asentados-en-el-registro-federal-refes"

def text(v): return "" if v is None else str(v).strip()
def canon(v):
    s=unicodedata.normalize("NFKD",text(v)); s="".join(c for c in s if not unicodedata.combining(c)).lower()
    return re.sub(r"[^a-z0-9]+"," ",s).strip()
def norm(v): return re.sub(r"\s+"," ",canon(v)).strip()
def num(v):
    s=text(v).replace(" ","")
    if not s: return None
    if s.count(",")==1 and s.count(".")==0: s=s.replace(",",".")
    try: return float(s)
    except ValueError: return None
def pg_array(values):
    out=[]
    for v in values:
        v=text(v)
        if v and v not in out: out.append(v)
    return "{"+",".join('"'+v.replace("\\","\\\\").replace('"','\\"')+'"' for v in out)+"}"
def field(row, *candidates):
    index={canon(k):v for k,v in row.items() if k is not None}
    wanted=[canon(v) for v in candidates]
    for key in wanted:
        if key in index and text(index[key]): return index[key]
    for key,value in index.items():
        if any(w and (key.startswith(w) or w in key) for w in wanted) and text(value): return value
    return None
def coordinates(row):
    lat=num(field(row,"latitud","latitude","lat")); lng=num(field(row,"longitud","longitude","lon","lng"))
    pairs=[]
    if lat is not None and lng is not None: pairs.append((lat,lng)); pairs.append((lng,lat))
    combined=text(field(row,"coordenadas","coordenada","gps","georreferencia","geolocalizacion"))
    vals=[float(v.replace(",",".")) for v in re.findall(r"-?\d+(?:[\.,]\d+)?",combined)] if combined else []
    if len(vals)>=2: pairs.extend([(vals[0],vals[1]),(vals[1],vals[0])])
    for a,b in pairs:
        if -56.5<=a<=-21 and -74.5<=b<=-52: return a,b
    return None
def classify(row,name):
    type_text=" ".join(text(field(row,k)) for k in ["tipologia","tipo establecimiento","categoria","especialidad","dependencia","nivel atencion","nombre tipologia"])
    blob=canon(name+" "+type_text); caps=[]
    def add(x):
        if x not in caps: caps.append(x)
    if re.search(r"hospital|internacion|sanatorio",blob): add("hospital")
    if re.search(r"odont|dental",blob): add("dental")
    if re.search(r"laborator|bioquim|analisis clin",blob): add("lab")
    if re.search(r"radiolog|diagnostico por imagen|tomograf|resonancia|ecograf",blob): add("imaging")
    if re.search(r"medicina laboral|salud ocupacional|medicina del trabajo",blob): add("occupational_health_clinic")
    if re.search(r"centro de salud|atencion primaria|aps|consultorio|unidad sanitaria|puesto sanitario|centro medico",blob): add("general_practitioner")
    specialty=text(field(row,"especialidad","especialidades","servicios"))
    if specialty: add("specialist")
    if not caps: add("healthcare_facility")
    priority=["occupational_health_clinic","dental","lab","imaging","hospital","general_practitioner","specialist","healthcare_facility"]
    return next((x for x in priority if x in caps),"healthcare_facility"),caps+[specialty] if specialty and specialty not in caps else caps
def location_key(name,address,lat,lng):
    payload=json.dumps({"name":norm(name),"address":address.lower(),"country":"AR","lat":round(lat,6),"lng":round(lng,6)},sort_keys=True,ensure_ascii=False,separators=(",",":"))
    return "loc:"+hashlib.sha256(payload.encode()).hexdigest()
def normalize(row):
    c=coordinates(row)
    if not c: return None
    lat,lng=c
    name=text(field(row,"nombre establecimiento","establecimiento","nombre efector","nombre","razon social"))
    if not name: return None
    code=text(field(row,"codigo federal","codigo establecimiento","codigo refes","id refes","refes","codigo","id establecimiento","id"))
    if not code: code=hashlib.sha1(f"{name}|{lat:.6f}|{lng:.6f}".encode()).hexdigest()[:20]
    address=text(field(row,"domicilio","direccion","calle","direccion establecimiento")); number=text(field(row,"numero","altura","numero puerta"))
    if address and number and number not in address: address=f"{address} {number}".strip()
    city=text(field(row,"localidad","ciudad","municipio")); department=text(field(row,"departamento","partido")); province=text(field(row,"provincia")); postal=text(field(row,"codigo postal","cp"))
    full=", ".join(v for v in [address,postal,city,department,province,"Argentina"] if v)
    primary,caps=classify(row,name)
    return [f"refes:{code}",DATASET_URL,name,norm(name),address,full,city,province,postal,"AR",f"{lat:.8f}",f"{lng:.8f}",text(field(row,"telefono","telefonos","tel")),text(field(row,"sitio web","website","web","url")),text(field(row,"email","correo electronico","correo")).lower(),primary,pg_array(caps),"0.99" if address else "0.96",location_key(name,full,lat,lng)]

def read_rows(path):
    suffix=Path(path).suffix.lower()
    if suffix in [".xlsx",".xlsm"]:
        from openpyxl import load_workbook
        wb=load_workbook(path,read_only=True,data_only=True); ws=wb.active
        rows=ws.iter_rows(values_only=True); probe=[]
        for _ in range(25):
            try: probe.append(next(rows))
            except StopIteration: break
        best_i=0; best_score=-1
        for i,r in enumerate(probe):
            keys=[canon(v) for v in r if text(v)]
            score=sum(3 for k in keys if "establecimiento" in k)+sum(2 for k in keys if "latitud" in k or "longitud" in k)+sum(1 for k in keys if "codigo" in k or "provincia" in k)
            if score>best_score: best_i=i; best_score=score
        headers=[text(v) or f"column_{i}" for i,v in enumerate(probe[best_i])]
        for r in probe[best_i+1:]: yield dict(zip(headers,r))
        for r in rows: yield dict(zip(headers,r))
    else:
        with open(path,"r",encoding="utf-8-sig",newline="") as f:
            sample=f.read(131072); f.seek(0)
            try: dialect=csv.Sniffer().sniff(sample,delimiters=";,\t|")
            except csv.Error: dialect=csv.excel; dialect.delimiter="," 
            yield from csv.DictReader(f,dialect=dialect)

def main():
    p=argparse.ArgumentParser(); p.add_argument("--input",required=True); p.add_argument("--output",required=True); a=p.parse_args()
    with open(a.output,"w",encoding="utf-8",newline="") as out:
        writer=csv.writer(out,delimiter="\t",quoting=csv.QUOTE_ALL,lineterminator="\n"); writer.writerow(COLUMNS)
        seen=set(); n=ok=bad=0
        for r in read_rows(a.input):
            n+=1; row=normalize(r)
            if not row or row[0] in seen: bad+=1; continue
            seen.add(row[0]); writer.writerow(row); ok+=1
    print(json.dumps({"source":"ar_refes","inputRows":n,"outputRows":ok,"rejectedRows":bad}))
if __name__=="__main__": main()
