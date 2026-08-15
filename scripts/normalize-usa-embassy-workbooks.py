#!/usr/bin/env python3

"""Normalize the final geocoded U.S. Embassy provider workbooks into the canonical staging TSV."""

from __future__ import annotations

import argparse
import csv
import hashlib
import math
import re
from pathlib import Path

from openpyxl import load_workbook

OUTPUT_COLUMNS = [
    "source_record_id",
    "source_url",
    "name",
    "normalized_name",
    "address_line1",
    "formatted_address",
    "city",
    "state_region",
    "postal_code",
    "country_code",
    "lat",
    "lng",
    "phone",
    "website",
    "email",
    "primary_provider_type",
    "capability_tags",
    "quality_score",
    "master_key",
]

CANONICAL_TYPES = {
    "urgent_care",
    "dot_provider",
    "faa_provider",
    "lab",
    "general_practitioner",
    "occupational_health_clinic",
    "dental",
    "imaging",
    "pharmacy_vaccination",
    "hospital",
    "specialist",
    "unknown",
}

SOURCE_ID_RE = re.compile(r"^usembassy-\d{8}-\d{5}$")


def clean(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    # Keep the staging TSV one physical line per provider. Excel cells can contain
    # line breaks/tabs copied from embassy pages.
    return re.sub(r"[\t\r\n]+", " ", text).strip()


def truthy(value: object) -> bool:
    if isinstance(value, bool):
        return value
    return clean(value).lower() in {"1", "true", "yes", "y"}


def number(value: object, field: str, sequence: int) -> float:
    try:
        parsed = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Sequence {sequence}: invalid {field}={value!r}") from exc
    if not math.isfinite(parsed):
        raise ValueError(f"Sequence {sequence}: non-finite {field}={value!r}")
    return parsed


def pg_array_literal(value: object) -> str:
    raw = clean(value)
    if not raw:
        return "{}"
    # The final workbooks store capability tags as a semicolon-delimited display
    # string. Preserve every non-empty tag while producing PostgreSQL text[] input.
    tags: list[str] = []
    seen: set[str] = set()
    for item in raw.split(";"):
        tag = item.strip()
        if not tag or tag in seen:
            continue
        seen.add(tag)
        tags.append(tag)

    def quote(tag: str) -> str:
        return '"' + tag.replace("\\", "\\\\").replace('"', '\\"') + '"'

    return "{" + ",".join(quote(tag) for tag in tags) + "}"


def fallback_master_key(row: dict[str, object], sequence: int) -> str:
    seed = "|".join(
        [
            clean(row.get("normalized_name")) or clean(row.get("facility_name")).lower(),
            clean(row.get("country_code")).upper(),
            clean(row.get("latitude")),
            clean(row.get("longitude")),
            str(sequence),
        ]
    )
    return "loc:" + hashlib.sha256(seed.encode("utf-8")).hexdigest()


def rows_from_workbook(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Providers" not in workbook.sheetnames:
        raise ValueError(f"{path.name}: missing Providers sheet")
    sheet = workbook["Providers"]
    iterator = sheet.iter_rows(values_only=True)
    try:
        header_values = next(iterator)
    except StopIteration as exc:
        raise ValueError(f"{path.name}: Providers sheet is empty") from exc
    headers = [clean(value) for value in header_values]
    index = {name: position for position, name in enumerate(headers)}
    required = {
        "sequence",
        "source_record_id",
        "facility_name",
        "latitude",
        "longitude",
        "valid_coordinates",
        "country_code",
        "normalized_name",
        "primary_provider_type",
        "capability_tags",
        "quality_score",
        "master_key",
    }
    missing = sorted(required - set(index))
    if missing:
        raise ValueError(f"{path.name}: missing required columns: {', '.join(missing)}")

    for values in iterator:
        if not any(value is not None and clean(value) for value in values):
            continue
        row = {name: values[position] if position < len(values) else None for name, position in index.items()}
        yield row


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", required=True)
    parser.add_argument("--expected", type=int, default=14552)
    parser.add_argument("workbooks", nargs="+")
    args = parser.parse_args()

    if len(args.workbooks) != 3:
        raise SystemExit("Exactly three final Embassy workbooks are required")

    records: dict[int, list[str]] = {}
    source_ids: set[str] = set()
    master_keys: set[str] = set()
    country_counts: dict[str, int] = {}
    type_counts: dict[str, int] = {}
    fallback_master_keys = 0

    for workbook_argument in args.workbooks:
        workbook_path = Path(workbook_argument).resolve()
        if not workbook_path.is_file():
            raise ValueError(f"Missing workbook: {workbook_path}")
        for row in rows_from_workbook(workbook_path):
            try:
                sequence = int(float(row["sequence"]))
            except (TypeError, ValueError) as exc:
                raise ValueError(f"{workbook_path.name}: invalid sequence {row.get('sequence')!r}") from exc
            if sequence in records:
                raise ValueError(f"Duplicate sequence {sequence}")

            source_record_id = clean(row.get("source_record_id"))
            if not SOURCE_ID_RE.match(source_record_id):
                raise ValueError(f"Sequence {sequence}: invalid source_record_id {source_record_id!r}")
            if source_record_id in source_ids:
                raise ValueError(f"Duplicate source_record_id {source_record_id}")
            source_ids.add(source_record_id)

            name = clean(row.get("facility_name"))
            if not name:
                raise ValueError(f"Sequence {sequence}: provider name is blank")
            if not truthy(row.get("valid_coordinates")):
                raise ValueError(f"Sequence {sequence}: coordinates are not marked valid")
            lat = number(row.get("latitude"), "latitude", sequence)
            lng = number(row.get("longitude"), "longitude", sequence)
            if not (-90 <= lat <= 90 and -180 <= lng <= 180):
                raise ValueError(f"Sequence {sequence}: coordinates out of range")

            country_code = clean(row.get("country_code")).upper() or "XX"
            primary_type = clean(row.get("primary_provider_type")).lower() or "unknown"
            if primary_type not in CANONICAL_TYPES:
                raise ValueError(f"Sequence {sequence}: non-canonical provider type {primary_type!r}")

            quality_raw = row.get("quality_score")
            quality = number(quality_raw, "quality_score", sequence) if clean(quality_raw) else 0.5
            quality = max(0.0, min(1.0, quality))

            master_key = clean(row.get("master_key"))
            if not master_key:
                master_key = fallback_master_key(row, sequence)
                fallback_master_keys += 1
            master_keys.add(master_key)

            website = clean(row.get("website"))
            phone = clean(row.get("international_phone")) or clean(row.get("local_phone"))
            formatted_address = clean(row.get("formatted_address")) or clean(row.get("original_address"))
            city = clean(row.get("geocoded_city")) or clean(row.get("original_city"))
            state_region = clean(row.get("state_region")) or clean(row.get("region"))

            records[sequence] = [
                source_record_id,
                "",  # source page URL was not retained in the final workbook; do not invent one
                name,
                clean(row.get("normalized_name")),
                clean(row.get("original_address")),
                formatted_address,
                city,
                state_region,
                clean(row.get("postal_code")),
                country_code,
                format(lat, ".15g"),
                format(lng, ".15g"),
                phone,
                website,
                clean(row.get("email")),
                primary_type,
                pg_array_literal(row.get("capability_tags")),
                format(quality, ".6f"),
                master_key,
            ]
            country_counts[country_code] = country_counts.get(country_code, 0) + 1
            type_counts[primary_type] = type_counts.get(primary_type, 0) + 1

    expected_sequences = set(range(1, args.expected + 1))
    actual_sequences = set(records)
    if actual_sequences != expected_sequences:
        missing = sorted(expected_sequences - actual_sequences)[:20]
        extra = sorted(actual_sequences - expected_sequences)[:20]
        raise ValueError(
            f"Sequence coverage mismatch: rows={len(records)} expected={args.expected} "
            f"missing={missing} extra={extra}"
        )

    output_path = Path(args.output).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t", quotechar='"', quoting=csv.QUOTE_MINIMAL, lineterminator="\n")
        writer.writerow(OUTPUT_COLUMNS)
        for sequence in range(1, args.expected + 1):
            writer.writerow(records[sequence])

    print(
        {
            "rows": len(records),
            "distinct_source_record_ids": len(source_ids),
            "distinct_master_keys": len(master_keys),
            "countries": len(country_counts),
            "unknown_country_rows": country_counts.get("XX", 0),
            "fallback_master_keys": fallback_master_keys,
            "provider_types": dict(sorted(type_counts.items())),
            "output": str(output_path),
        }
    )


if __name__ == "__main__":
    main()
