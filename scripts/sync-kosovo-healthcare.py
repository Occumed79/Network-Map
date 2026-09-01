#!/usr/bin/env python3

import argparse
import csv
import hashlib
import json
import math
import re
import time
import unicodedata
from pathlib import Path

import requests
from openpyxl import load_workbook

SOURCE_URL = "https://msh.rks-gov.net/Department/GetDocument?fileName=40874433.7895.xlsx&original=Lista+e+Institucioneve+Private+Sh%C3%ABndet%C3%ABsore+t%C3%AB+licencuara+2021-2026.xlsx"
DATASET_URL = "https://msh.rks-gov.net/Department/Index/1060?type=1"
NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
USER_AGENT = "Occu-Med-Network-Map/1.0 (+https://github.com/Occumed79/Network-Map)"

COLUMNS = [
    "source_record_id", "source_url", "name", "normalized_name", "address_line1",
    "formatted_address", "city", "state_region", "postal_code", "country_code",
    "lat", "lng", "phone", "website", "email", "primary_provider_type",
    "capability_tags", "quality_score", "master_key",
]


def text(value):
    return "" if value is None else str(value).strip()


def ascii_key(value):
    normalized = unicodedata.normalize("NFKD", text(value))
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    normalized = normalized.lower().replace("ë", "e").replace("ç", "c")
    return re.sub(r"[^a-z0-9]+", " ", normalized).strip()


def normalized_name(value):
    return re.sub(r"\s+", " ", ascii_key(value)).strip()


def hash_text(value):
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def valid_coordinates(lat, lng):
    return (
        isinstance(lat, (int, float)) and isinstance(lng, (int, float))
        and math.isfinite(lat) and math.isfinite(lng)
        and 41.7 <= lat <= 43.4 and 19.8 <= lng <= 23.0
    )


def load_existing(path):
    result = {}
    if not path or not Path(path).exists():
        return result
    with open(path, "r", encoding="utf-8") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            try:
                lat = float(row.get("lat") or "")
                lng = float(row.get("lng") or "")
            except ValueError:
                continue
            if valid_coordinates(lat, lng):
                result[text(row.get("source_record_id"))] = (lat, lng)
    return result


def download_xlsx(path):
    response = requests.get(
        SOURCE_URL,
        headers={"User-Agent": USER_AGENT, "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*"},
        timeout=180,
    )
    response.raise_for_status()
    if len(response.content) < 5000:
        raise RuntimeError("Kosovo Ministry workbook download was unexpectedly small")
    Path(path).write_bytes(response.content)


def find_header(rows):
    signals = ("institucion", "emri", "adresa", "komuna", "licenc", "veprimtar", "sherb")
    best = None
    best_score = -1
    for idx, row in enumerate(rows[:40]):
        keys = [ascii_key(cell) for cell in row]
        joined = " | ".join(keys)
        score = sum(1 for signal in signals if signal in joined)
        if score > best_score and sum(bool(key) for key in keys) >= 2:
            best = idx
            best_score = score
    if best is None or best_score < 2:
        raise RuntimeError("Could not identify Kosovo workbook header row")
    return best


def workbook_rows(path):
    book = load_workbook(path, read_only=True, data_only=True)
    all_rows = []
    for sheet in book.worksheets:
        rows = [[text(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
        if not rows:
            continue
        try:
            header_idx = find_header(rows)
        except RuntimeError:
            continue
        headers = [ascii_key(value) or f"column_{idx}" for idx, value in enumerate(rows[header_idx])]
        for raw in rows[header_idx + 1:]:
            if not any(text(value) for value in raw):
                continue
            row = {headers[idx]: text(raw[idx]) if idx < len(raw) else "" for idx in range(len(headers))}
            row["__sheet"] = sheet.title
            all_rows.append(row)
    if not all_rows:
        raise RuntimeError("Kosovo workbook yielded no data rows")
    return all_rows


def pick(row, exact=(), contains=()):
    for key in exact:
        value = text(row.get(ascii_key(key)))
        if value:
            return value
    for key, value in row.items():
        normalized = ascii_key(key)
        if any(token in normalized for token in contains) and text(value):
            return text(value)
    return ""


def classify(name, kind, services):
    haystack = ascii_key(" ".join([name, kind, services]))
    primary = "healthcare_facility"
    if re.search(r"dent|stomat|odont", haystack):
        primary = "dental"
    elif re.search(r"laborator|biokimi|mikrobiolog|patolog", haystack):
        primary = "lab"
    elif re.search(r"radiolog|imazheri|imaging|rentgen|rreze|ultraz|mri|ct scan", haystack):
        primary = "imaging"
    elif re.search(r"spital|hospital", haystack):
        primary = "hospital"
    elif re.search(r"mjekesi familj|family medicine|ambulant|primary|qender mjek", haystack):
        primary = "general_practitioner"
    elif re.search(r"kardiolog|pneumolog|neurolog|psikiatr|ortoped|gjinekolog|oftalmolog|special", haystack):
        primary = "specialist"
    tags = [primary, "healthcare_facility", "kosovo_moh_licensed_private"] if primary != "healthcare_facility" else ["healthcare_facility", "kosovo_moh_licensed_private"]
    return primary, list(dict.fromkeys(tags))


def geocode(session, query):
    response = session.get(
        NOMINATIM_URL,
        params={"format": "jsonv2", "limit": 1, "countrycodes": "xk", "q": query},
        headers={"User-Agent": USER_AGENT, "Accept-Language": "sq,en"},
        timeout=45,
    )
    if response.status_code == 429:
        time.sleep(3)
        return None
    response.raise_for_status()
    payload = response.json()
    if not payload:
        return None
    try:
        lat = float(payload[0]["lat"])
        lng = float(payload[0]["lon"])
    except (KeyError, TypeError, ValueError):
        return None
    return (lat, lng) if valid_coordinates(lat, lng) else None


def pg_array(values):
    escaped = []
    for value in values:
        value = str(value).replace("\\", "\\\\").replace('"', '\\"')
        escaped.append(f'"{value}"')
    return "{" + ",".join(escaped) + "}"


def build_rows(records, existing):
    output = {}
    session = requests.Session()
    last_geocode = 0.0
    geocoded = 0
    reused = 0
    skipped = 0

    for row in records:
        name = pick(row,
            exact=("emri i institucionit", "emri i institucionit shendetesor", "emri"),
            contains=("emri i institucion", "institucion shendetes", "institucioni"),
        )
        if not name or len(name) < 3:
            continue
        address = pick(row, exact=("adresa", "adresa e institucionit"), contains=("adresa", "lokacion"))
        municipality = pick(row, exact=("komuna",), contains=("komuna", "municip"))
        license_no = pick(row,
            exact=("nr i licences", "numri i licences", "nr licences", "licenca"),
            contains=("licenc", "license"),
        )
        kind = pick(row, exact=("lloji i institucionit", "lloji"), contains=("lloji", "tipi"))
        services = pick(row, exact=("veprimtaria", "sherbimet"), contains=("veprimtar", "sherb"))
        phone = pick(row, contains=("telefon", "phone", "tel "))
        email = pick(row, contains=("email", "e mail"))

        identity = license_no or hash_text("|".join([normalized_name(name), ascii_key(address), ascii_key(municipality)]))[:20]
        source_id = f"xk-moh:{identity}"
        coords = existing.get(source_id)
        if coords:
            reused += 1
        else:
            query = ", ".join(part for part in [name, address, municipality, "Kosovo"] if part)
            if not address and not municipality:
                skipped += 1
                continue
            wait = 1.10 - (time.monotonic() - last_geocode)
            if wait > 0:
                time.sleep(wait)
            try:
                coords = geocode(session, query)
            except requests.RequestException:
                coords = None
            last_geocode = time.monotonic()
            if not coords and address and municipality:
                wait = 1.10 - (time.monotonic() - last_geocode)
                if wait > 0:
                    time.sleep(wait)
                try:
                    coords = geocode(session, f"{address}, {municipality}, Kosovo")
                except requests.RequestException:
                    coords = None
                last_geocode = time.monotonic()
            if not coords:
                skipped += 1
                continue
            geocoded += 1

        lat, lng = coords
        primary, tags = classify(name, kind, services)
        formatted = ", ".join(part for part in [address, municipality, "Kosovo"] if part)
        normalized = normalized_name(name)
        master_key = "loc:" + hash_text(json.dumps({
            "name": normalized,
            "address": formatted.lower(),
            "country": "XK",
            "lat": round(lat, 6),
            "lng": round(lng, 6),
        }, sort_keys=True, ensure_ascii=False))

        output[source_id] = [
            source_id, DATASET_URL, name, normalized, address, formatted, municipality, municipality, "", "XK",
            lat, lng, phone, "", email, primary, pg_array(tags), 0.96, master_key,
        ]

    print(json.dumps({
        "source": "xk_moh_private_licensed",
        "scanned": len(records),
        "map_rows": len(output),
        "coordinates_reused": reused,
        "newly_geocoded": geocoded,
        "skipped_unplaced": skipped,
    }, ensure_ascii=False))
    return list(output.values())


def write_tsv(path, rows):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t", quoting=csv.QUOTE_ALL, lineterminator="\n")
        writer.writerow(COLUMNS)
        writer.writerows(sorted(rows, key=lambda row: (str(row[2]).lower(), str(row[0]))))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", required=True)
    parser.add_argument("--existing", default="")
    parser.add_argument("--workbook", default="/tmp/kosovo-healthcare.xlsx")
    args = parser.parse_args()

    download_xlsx(args.workbook)
    records = workbook_rows(args.workbook)
    rows = build_rows(records, load_existing(args.existing))
    if not rows:
        raise RuntimeError("Kosovo normalization produced zero map-renderable licensed institutions")
    write_tsv(args.output, rows)


if __name__ == "__main__":
    main()
